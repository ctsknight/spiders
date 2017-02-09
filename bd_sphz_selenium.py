# -*- coding: utf-8 -*-
from pyvirtualdisplay import Display
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import random
import re



rare_words_xls = '1.xlsx'
rare_words = {}


def search_rare_words():
    print 'begin to search rare words...'
    display = Display(visible=0, size=(1280, 800))
    display.start()
    driver = webdriver.Firefox()
    #driver.get("http://www.google.com")
    #driver.get("https://www.baidu.com")
    #inputElement = driver.find_element_by_name("wd")

    search_rare_words_loop(driver, 0)

    driver.quit()
    display.stop()
    print 'finish searching rare words...'


def search_rare_words_loop(driver, deep):
    deep += 1
    'search rare words loop %d' % deep
    key = ''
    i = 1
    all_done = True
    failed_num = 0
    for (k, v) in rare_words.items():
        key = k
        if v == 0 or v == 'error':
            all_done = False
            print 'begin search %d word, failed_num = %d' % (i, failed_num)
            try:
                driver.get("https://www.baidu.com")
                inputElement = driver.find_element_by_name("wd")
                #print key
                #rare_words[key] = '11111'
                inputElement.clear()
                #key = k.encode('utf-8')
                #print key
                inputElement.send_keys(key)
                inputElement.submit()
            except:
                print 'error to find wd'
                failed_num += 1
                i += 1
                continue
            try:
                # we have to wait for the page to refresh, the last thing that seems to be updated is the title
                #WebDriverWait(driver, 60).until(EC.title_contains(key))
                #print key
                WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//*[@id="1"]')
                    )
                )

                #
                #print driver.title
                #WebDriverWait(driver, 60).until(EC.title_contains(key))
                search_result_nums = driver.find_element_by_xpath('//div[@class="nums"]').text
                #print search_result_nums
                result = re.match(u'.+约(.+)个', search_result_nums, re.S)
                #print result.group(1)
                #result = re.match('About (.+) results', search_result_nums)
                #print result.group(1)
                rare_words[key] = result.group(1)
                print 'finished search %d word, failed_num = %d' % (i, failed_num)
            except:
                print 'error'
                rare_words[key] = 'error'
                failed_num += 1
            i += 1
        
        #driver.implicitly_wait(random.randint(1, 4))
    #if all_done:
    #    return
    #else:
    #    search_rare_words_loop(driver, deep)


def read_rare_words_from_xls(filename):
    print 'begin to read xls...'
    wb = load_workbook(filename=filename)
    sheet = wb.worksheets[0]
    i = 1
    while(sheet['A' + str(i)].value is not None):
        if sheet['B' + str(i)].value is not None:
            rare_words.setdefault(sheet['A' + str(i)].value, sheet['B' + str(i)].value)
        else:
            rare_words.setdefault(sheet['A' + str(i)].value, 0)
        i += 1

    print 'finish reading xls...'

    i = 0
    j = 0
    for (k, v) in rare_words.items():
        if v == 0 or v == 'error':
            i += 1
        else:
            j += 1
        #k = k.encode('utf-8')
        #print k

    print 'error = %d, and finished= %d' % (i, j)


def write_sreach_result_to_xls(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.worksheets[0]
    i = 1
    while(sheet['A' + str(i)].value is not None):
        k = sheet['A' + str(i)].value
        if rare_words[k] is not None:
            sheet['B' + str(i)].value = rare_words[k]
            #print 'i=%d,and value=%s' %(i, str(rare_words[k]))
        i += 1
    wb.save("1.xlsx")

if __name__ == '__main__':
    read_rare_words_from_xls(rare_words_xls)
    search_rare_words()
    write_sreach_result_to_xls(rare_words_xls)
